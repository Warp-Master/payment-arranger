import argparse
import logging
import os
import re
import shutil
import warnings
from pathlib import Path

import openpyxl



def dir_path(path) -> Path:
    if os.path.isdir(path):
        return Path(path)
    else:
        raise argparse.ArgumentTypeError(f"readable_dir:{path} is not a valid path")


def build_argparser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-n",
        "--dry-run",
        dest="is_dry_run",
        action="store_true",
        help="disable file movement and subdirectory creation",
    )
    parser.add_argument(
        "-l", "--log", help="specify log file", type=Path, default="payment.log"
    )
    parser.add_argument(
        "-i",
        "--inp",
        type=dir_path,
        default=Path("."),
        help="directory with .xlsx files",
    )
    parser.add_argument(
        "-d",
        "--dst",
        type=Path,
        default=Path("."),
        help="directory where subdirectories with files will be created",
    )
    return parser


def configure_logging(args) -> None:
    # disable warnings about style from openpyxl
    warnings.filterwarnings(
        "ignore", category=UserWarning, module=re.escape("openpyxl.styles.stylesheet")
    )

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[logging.FileHandler(args.log), logging.StreamHandler()],
    )


def main(args=None) -> None:
    argparser = build_argparser()
    args = argparser.parse_args(args)
    configure_logging(args)
    for filename in args.inp.glob("*.xlsx"):
        # read payment info
        wb = openpyxl.load_workbook(filename, data_only=True, read_only=True)
        ws = wb.active
        inn = ws["D21"].value
        reciever = ws["B22"].value
        purpose = ws["B27"].value
        amount = ws["T11"].value
        logging.info(f"[payment] to={reciever} for={purpose} ({amount})")
        wb.close()

        if not args.is_dry_run:
            # move file to new location
            dir = args.dst / inn
            os.makedirs(dir, exist_ok=True)
            shutil.move(filename, dir / filename.name)


if __name__ == "__main__":
    main()
